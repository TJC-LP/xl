#!/usr/bin/env python3
"""Generate the synthetic real-file fixture corpus for xl-ooxml/xl-cats-effect tests.

Produces xl-ooxml/test/resources/fixtures/*.xlsx with openpyxl (a producer whose
XML dialect differs from XL's own writer), then converts a subset through
LibreOffice headless to capture a third producer dialect (*-lo.xlsx).

ALL content is synthetic/invented. No confidential data.

Determinism:
  - All cell dates/times are fixed literals (no `now()`, no randomness).
  - docProps timestamps are pinned to FIXED_STAMP.
  - Generated zips are post-processed so every entry carries a fixed timestamp;
    regenerating an openpyxl fixture on the same machine is byte-stable.
  - LibreOffice embeds its own generation metadata, so *-lo.xlsx files are NOT
    byte-stable across regenerations; the committed files are canonical.
    Pass --skip-lo to leave the committed *-lo.xlsx files untouched.

Usage:
  python3 scripts/generate-fixtures.py            # regenerate everything
  python3 scripts/generate-fixtures.py --skip-lo  # skip LibreOffice variants

Requires: python3, openpyxl (3.1.x). LibreOffice (soffice) optional for -lo files.
"""

from __future__ import annotations

import argparse
import base64
import datetime as dt
import re
import shutil
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path

import openpyxl
from openpyxl.chart import BarChart, Reference, ScatterChart, Series as ChartSeries
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image
from openpyxl.formatting.rule import (
    CellIsRule,
    ColorScaleRule,
    DataBarRule,
    FormulaRule,
    IconSetRule,
    Rule,
)
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parent.parent
FIXTURES_DIR = REPO_ROOT / "xl-ooxml" / "test" / "resources" / "fixtures"

# Pinned docProps timestamp (UTC) - keeps docProps/core.xml stable.
FIXED_STAMP = dt.datetime(2024, 1, 15, 9, 30, 0)
# Pinned zip entry timestamp (zip format has no timezone).
ZIP_STAMP = (2024, 1, 15, 9, 30, 0)

# 3x3 solid-color RGB PNG (73 bytes), hand-assembled with zlib - no PIL needed.
# Construction: PNG signature + IHDR(3x3, 8-bit RGB) + IDAT(zlib level 9) + IEND.
TINY_PNG_B64 = (
    "iVBORw0KGgoAAAANSUhEUgAAAAMAAAADCAIAAADZSiLoAAAAEElEQVR42mN4UawDQQxYWADU"
    "6w3A6nptbAAAAABJRU5ErkJggg=="
)
TINY_PNG = base64.b64decode(TINY_PNG_B64)


def new_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.properties.created = FIXED_STAMP
    wb.properties.modified = FIXED_STAMP
    wb.properties.creator = "xl fixture generator (synthetic data)"
    return wb


def normalize_zip(path: Path) -> None:
    """Rewrite the zip with fixed entry timestamps and a pinned dcterms:modified.

    openpyxl overwrites docProps/core.xml's <dcterms:modified> with now() at
    save time, so byte-stable regeneration requires pinning it here.
    """
    with zipfile.ZipFile(path, "r") as zin:
        entries = [(info.filename, zin.read(info.filename)) for info in zin.infolist()]
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED, compresslevel=6) as zout:
        for name, data in entries:
            if name == "docProps/core.xml":
                data = re.sub(
                    rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
                    rb"\g<1>2024-01-15T09:30:00Z\g<2>",
                    data,
                )
            info = zipfile.ZipInfo(name, date_time=ZIP_STAMP)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = 0o600 << 16
            zout.writestr(info, data)


def save(wb: openpyxl.Workbook, name: str) -> Path:
    out = FIXTURES_DIR / name
    wb.save(out)
    normalize_zip(out)
    print(f"  wrote {out.relative_to(REPO_ROOT)} ({out.stat().st_size} bytes)")
    return out


def gen_small_values() -> Path:
    """Strings (unicode + whitespace-significant), numbers, booleans, dates."""
    wb = new_workbook()
    ws = wb.active
    ws.title = "Values"
    # Strings: plain, accented latin, CJK, astral-plane emoji,
    # leading/trailing whitespace, internal runs of spaces.
    ws["A1"] = "plain"
    ws["A2"] = "héllo wörld"
    ws["A3"] = "日本語テキスト"
    ws["A4"] = "emoji \U0001f680 rocket"
    ws["A5"] = "  leading and trailing  "
    ws["A6"] = "internal  double  spaces"
    # Numbers: int, negative, decimal, large, tiny (scientific-notation prone).
    ws["B1"] = 42
    ws["B2"] = -17
    ws["B3"] = 3.14159
    ws["B4"] = 1234567890123
    ws["B5"] = 1.23e-10
    ws["B6"] = 9.99e15
    # Booleans.
    ws["C1"] = True
    ws["C2"] = False
    # Dates (fixed; openpyxl stores serial number + date numFmt).
    ws["D1"] = dt.date(2024, 3, 15)
    ws["D2"] = dt.datetime(2024, 3, 15, 14, 30, 45)
    # Digit string (must stay text, not number).
    ws["E1"] = "0123"
    return save(wb, "small-values.xlsx")


def gen_styled() -> Path:
    """Fonts, fills, borders, numFmts, alignment + indent, merges."""
    wb = new_workbook()
    ws = wb.active
    ws.title = "Styled"
    ws["A1"] = "bold red on yellow"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="FFCC0000")
    ws["A1"].fill = PatternFill(fill_type="solid", start_color="FFFFFF00")
    thin = Side(style="thin", color="FF000000")
    ws["A2"] = "italic thin box"
    ws["A2"].font = Font(italic=True)
    ws["A2"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws["B1"] = 1234.5
    ws["B1"].number_format = "$#,##0.00"
    ws["B2"] = 0.4567
    ws["B2"].number_format = "0.00%"
    ws["B3"] = dt.date(2024, 6, 30)
    ws["B3"].number_format = "yyyy-mm-dd"
    ws["C1"] = "wrapped centered indented"
    ws["C1"].alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True, indent=2
    )
    ws["D1"] = "thick bottom"
    ws["D1"].border = Border(bottom=Side(style="thick", color="FF333333"))
    ws.merge_cells("A4:C4")
    ws["A4"] = "Merged Header"
    ws["A4"].alignment = Alignment(horizontal="center")
    return save(wb, "styled.xlsx")


def gen_formulas() -> Path:
    """Cross-sheet refs, ranges, leading-plus (=+ref), absolute anchors."""
    wb = new_workbook()
    data = wb.active
    data.title = "Data"
    for i, v in enumerate([10, 20, 30, 40, 50], start=1):
        data[f"A{i}"] = v
    for i, v in enumerate([1.5, 2.5, 3.5], start=1):
        data[f"B{i}"] = v
    calc = wb.create_sheet("Calc")
    calc["A1"] = "=SUM(Data!A1:A5)"
    calc["A2"] = "=Data!A1*2"
    calc["A3"] = "=+Data!A2"  # legacy Lotus leading plus
    calc["A4"] = "=$A$1+A2"  # absolute anchor
    calc["A5"] = "=SUM($A$1:A4)"  # mixed range anchor
    calc["A6"] = "=AVERAGE(Data!A1:A5)+Data!B1"
    return save(wb, "formulas.xlsx")


def gen_autofilter() -> Path:
    wb = new_workbook()
    ws = wb.active
    ws.title = "Filtered"
    rows = [
        ("Region", "Product", "Units"),
        ("North", "Anvil", 12),
        ("South", "Rocket", 7),
        ("East", "Magnet", 31),
        ("West", "Spring", 4),
        ("North", "Rocket", 19),
    ]
    for r, row in enumerate(rows, start=1):
        for c, v in enumerate(row, start=1):
            ws[f"{get_column_letter(c)}{r}"] = v
    ws.auto_filter.ref = "A1:C6"
    return save(wb, "autofilter.xlsx")


def gen_chart_bar() -> Path:
    """BarChart anchored next to its data (xl/charts/* + xl/drawings/*)."""
    wb = new_workbook()
    ws = wb.active
    ws.title = "ChartData"
    rows = [("Quarter", "Units"), ("Q1", 12), ("Q2", 19), ("Q3", 7), ("Q4", 23)]
    for r, row in enumerate(rows, start=1):
        ws[f"A{r}"], ws[f"B{r}"] = row
    chart = BarChart()
    chart.title = "Synthetic Units by Quarter"
    chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=5), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=5))
    ws.add_chart(chart, "D2")
    return save(wb, "chart-bar.xlsx")


def gen_chart_stacked() -> Path:
    """Stacked BarChart WITH overlap=100 (GH-222: pins the stacked read+write dialect).

    openpyxl does not set overlap automatically; without it Excel renders stacked
    bars side-by-side, so xl's reader requires overlap=100 for stacked groupings.
    """
    wb = new_workbook()
    ws = wb.active
    ws.title = "StackData"
    rows = [("Quarter", "North", "South"), ("Q1", 3, 7), ("Q2", 5, 2), ("Q3", 8, 4)]
    for r, row in enumerate(rows, start=1):
        ws[f"A{r}"], ws[f"B{r}"], ws[f"C{r}"] = row
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.title = "Synthetic Stacked Units"
    chart.add_data(
        Reference(ws, min_col=2, max_col=3, min_row=1, max_row=4), titles_from_data=True
    )
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=4))
    ws.add_chart(chart, "E2")
    return save(wb, "chart-stacked.xlsx")


def gen_chart_scatter() -> Path:
    """ScatterChart (GH-222: outside the typed fence — whole anchor stays Preserved)."""
    wb = new_workbook()
    ws = wb.active
    ws.title = "ScatterData"
    rows = [("X", "Y"), (1, 2), (2, 5), (3, 4), (4, 9)]
    for r, row in enumerate(rows, start=1):
        ws[f"A{r}"], ws[f"B{r}"] = row
    chart = ScatterChart()
    chart.title = "Synthetic Scatter"
    xs = Reference(ws, min_col=1, min_row=2, max_row=5)
    ys = Reference(ws, min_col=2, min_row=1, max_row=5)
    chart.series.append(ChartSeries(ys, xs, title_from_data=True))
    ws.add_chart(chart, "D2")
    return save(wb, "chart-scatter.xlsx")


def gen_image() -> Path:
    """Embedded PNG (xl/media/* + xl/drawings/*) without requiring Pillow.

    openpyxl's Image.__init__ insists on PIL just to sniff width/height, so we
    construct the instance manually and feed the writer raw PNG bytes.
    """
    wb = new_workbook()
    ws = wb.active
    ws.title = "HasImage"
    ws["A1"] = "tiny synthetic png anchored at B2"
    img = Image.__new__(Image)
    img.ref = None
    img.width = 3
    img.height = 3
    img.format = "png"
    img.anchor = "B2"
    img._data = lambda: TINY_PNG  # writer calls _data() for xl/media payload
    ws.add_image(img)
    return save(wb, "image.xlsx")


def gen_comments_hyperlinks() -> Path:
    wb = new_workbook()
    ws = wb.active
    ws.title = "Notes"
    ws["A1"] = "commented cell"
    ws["A1"].comment = Comment("Synthetic note: invented for tests.", "Fixture Bot")
    ws["B1"] = "example link"
    ws["B1"].hyperlink = "https://example.com/xl-fixtures"
    ws["B2"] = "internal link"
    ws["B2"].hyperlink = "#Notes!A1"
    return save(wb, "comments-hyperlinks.xlsx")


def gen_condformat() -> Path:
    """Conditional formatting (GH-136): the six typed rule families + iconSet (Preserved).

    Exercises both dxf fill dialects (openpyxl writes patternType="solid" with
    fgColor+bgColor), stopIfTrue, a 3-point colorScale, a plain dataBar, top10,
    containsText with openpyxl's canonical SEARCH formula, a multi-range sqref,
    and an iconSet that must ride through xl's CfRule.Preserved verbatim.
    """
    wb = new_workbook()
    ws = wb.active
    ws.title = "CondFmt"
    for i, v in enumerate([12, 250, 87, 101, 3, 999, 45, 150], start=2):
        ws[f"B{i}"] = v
        ws[f"C{i}"] = v * 2
        ws[f"D{i}"] = v % 97
        ws[f"F{i}"] = v + 1
        ws[f"G{i}"] = v - 1
    for i, t in enumerate(
        ["todo: ship", "done", "todo later", "ok", "todone", "x", "todo", "none"], start=2
    ):
        ws[f"E{i}"] = t
    for i, v in enumerate([5, 10, 15, 20], start=2):
        ws[f"H{i}"] = v
        ws[f"J{i}"] = v * 3

    red_text = Font(color="FF9C0006", bold=True)
    pink = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
    yellow = PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid")

    # 1+2. Two rules on the SAME range: cellIs+dxf, then expression with stopIfTrue.
    ws.conditional_formatting.add(
        "B2:B9",
        CellIsRule(operator="greaterThan", formula=["100"], fill=pink, font=red_text),
    )
    ws.conditional_formatting.add(
        "B2:B9",
        FormulaRule(formula=["$B2>AVERAGE($B$2:$B$9)"], stopIfTrue=True, fill=yellow),
    )
    # 3. 3-point color scale.
    ws.conditional_formatting.add(
        "C2:C9",
        ColorScaleRule(
            start_type="min",
            start_color="FFF8696B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFFFEB84",
            end_type="max",
            end_color="FF63BE7B",
        ),
    )
    # 4. Data bar (no minLength/maxLength so it stays inside xl's typed fence).
    ws.conditional_formatting.add(
        "D2:D9", DataBarRule(start_type="min", end_type="max", color="FF638EC6")
    )
    # 5. containsText with the canonical SEARCH formula against the range's top-left.
    ws.conditional_formatting.add(
        "E2:E9",
        Rule(
            type="containsText",
            operator="containsText",
            text="todo",
            formula=['NOT(ISERROR(SEARCH("todo",E2)))'],
            dxf=DifferentialStyle(font=red_text, fill=pink),
        ),
    )
    # 6. top10 (rank 3) with a dxf.
    ws.conditional_formatting.add(
        "F2:F9", Rule(type="top10", rank=3, dxf=DifferentialStyle(fill=yellow))
    )
    # 7. iconSet: outside xl's typed fence - pins CfRule.Preserved verbatim round-trip.
    ws.conditional_formatting.add(
        "G2:G9", IconSetRule("3Arrows", "percent", [0, 33, 67])
    )
    # 8. Multi-range sqref (space-separated) with a between rule.
    ws.conditional_formatting.add(
        "H2:H5 J2:J5",
        CellIsRule(operator="between", formula=["8", "40"], fill=pink),
    )
    return save(wb, "condformat.xlsx")


def convert_with_libreoffice(sources: list[Path]) -> None:
    """Convert fixtures through LibreOffice headless -> *-lo.xlsx variants."""
    soffice = shutil.which("soffice") or "/usr/local/bin/soffice"
    if not Path(soffice).exists():
        print("  WARNING: soffice not found; keeping committed *-lo.xlsx files")
        return
    with tempfile.TemporaryDirectory(prefix="xl-fixtures-lo-") as tmp:
        profile = Path(tmp) / "profile"
        outdir = Path(tmp) / "out"
        outdir.mkdir()
        for src in sources:
            subprocess.run(
                [
                    soffice,
                    "--headless",
                    f"-env:UserInstallation=file://{profile}",
                    "--convert-to",
                    "xlsx",
                    "--outdir",
                    str(outdir),
                    str(src),
                ],
                check=True,
                capture_output=True,
                timeout=120,
            )
            converted = outdir / src.name
            if not converted.exists():
                raise RuntimeError(f"LibreOffice did not produce {converted}")
            dest = FIXTURES_DIR / f"{src.stem}-lo.xlsx"
            shutil.copyfile(converted, dest)
            converted.unlink()
            print(f"  wrote {dest.relative_to(REPO_ROOT)} ({dest.stat().st_size} bytes)")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--skip-lo", action="store_true", help="skip LibreOffice variants")
    args = parser.parse_args()

    FIXTURES_DIR.mkdir(parents=True, exist_ok=True)
    print(f"openpyxl {openpyxl.__version__} -> {FIXTURES_DIR.relative_to(REPO_ROOT)}")
    small = gen_small_values()
    styled = gen_styled()
    formulas = gen_formulas()
    gen_autofilter()
    gen_chart_bar()
    gen_chart_stacked()
    gen_chart_scatter()
    gen_image()
    gen_comments_hyperlinks()
    condformat = gen_condformat()
    if args.skip_lo:
        print("  --skip-lo: keeping committed *-lo.xlsx files")
    else:
        convert_with_libreoffice([small, styled, formulas, condformat])

    total = sum(f.stat().st_size for f in FIXTURES_DIR.glob("*.xlsx"))
    print(f"total corpus size: {total} bytes")
    if total >= 1_000_000:
        print("ERROR: corpus exceeds 1 MB budget", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
